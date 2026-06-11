"""
Streaming spreadsheet-to-Markdown conversion.

The MarkItDown spreadsheet converters load whole workbooks into pandas data
frames. This module keeps the large-file path local, bounded, and explicit.
"""

from __future__ import annotations

import csv
import os
import tempfile
import time
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time as datetime_time
from pathlib import Path
from typing import Iterable, Iterator, Sequence, TextIO

from charset_normalizer import from_bytes


SPREADSHEET_EXTENSIONS = {".csv", ".xls", ".xlsx"}


class SpreadsheetConversionError(Exception):
    """Raised when a spreadsheet cannot be converted."""


class SpreadsheetLimitError(SpreadsheetConversionError):
    """Raised when a spreadsheet exceeds a configured safety limit."""


@dataclass(frozen=True)
class SpreadsheetConversionOptions:
    preview_rows_per_sheet: int = 100
    preview_max_chars: int = 500_000
    max_sheets: int = 100
    max_columns: int = 256
    max_cell_chars: int = 2_000
    max_output_chars: int = 25 * 1024 * 1024
    timeout_seconds: int = 120
    max_xlsx_uncompressed_bytes: int = 512 * 1024 * 1024
    max_xlsx_compression_ratio: int = 200
    header_scan_rows: int = 25
    skip_empty_rows: bool = True

    def __post_init__(self) -> None:
        positive_fields = {
            "preview_rows_per_sheet": self.preview_rows_per_sheet,
            "preview_max_chars": self.preview_max_chars,
            "max_sheets": self.max_sheets,
            "max_columns": self.max_columns,
            "max_cell_chars": self.max_cell_chars,
            "max_output_chars": self.max_output_chars,
            "timeout_seconds": self.timeout_seconds,
            "max_xlsx_uncompressed_bytes": self.max_xlsx_uncompressed_bytes,
            "max_xlsx_compression_ratio": self.max_xlsx_compression_ratio,
            "header_scan_rows": self.header_scan_rows,
        }
        for name, value in positive_fields.items():
            if value < 1:
                raise ValueError(f"{name} must be at least 1")


@dataclass
class SpreadsheetConversionResult:
    markdown_length: int
    preview: str
    preview_truncated: bool
    sheets: int
    rows: int
    warnings: list[str] = field(default_factory=list)


@dataclass
class _ConversionState:
    started_at: float
    sheets: int = 0
    rows: int = 0
    warnings: list[str] = field(default_factory=list)
    row_width_warning_sheets: set[str] = field(default_factory=set)
    cell_truncated: bool = False


class _LimitedWriter:
    def __init__(self, output: TextIO, max_chars: int):
        self._output = output
        self._max_chars = max_chars
        self.chars = 0

    def write(self, text: str) -> None:
        next_size = self.chars + len(text)
        if next_size > self._max_chars:
            raise SpreadsheetLimitError(
                f"Spreadsheet Markdown output exceeded {self._max_chars} characters."
            )
        self._output.write(text)
        self.chars = next_size


class _PreviewBuilder:
    def __init__(self, max_chars: int):
        self._max_chars = max_chars
        self._parts: list[str] = []
        self.chars = 0
        self.truncated = False

    def write(self, text: str) -> None:
        if not text or self.chars >= self._max_chars:
            if text:
                self.truncated = True
            return

        remaining = self._max_chars - self.chars
        if len(text) > remaining:
            self._parts.append(text[:remaining].rstrip())
            self.chars = self._max_chars
            self.truncated = True
            return

        self._parts.append(text)
        self.chars += len(text)

    def add_marker(self, message: str) -> None:
        self.truncated = True
        self.write(f"\n\n> {message}\n\n")

    def text(self) -> str:
        return "".join(self._parts).rstrip()


def is_spreadsheet_path(path_or_name: str | Path) -> bool:
    return Path(path_or_name).suffix.lower() in SPREADSHEET_EXTENSIONS


def convert_spreadsheet_to_path(
    source: str | Path,
    target: str | Path,
    options: SpreadsheetConversionOptions | None = None,
    display_name: str | None = None,
) -> SpreadsheetConversionResult:
    """Convert a spreadsheet to Markdown using an atomic output write."""
    source_path = Path(source)
    target_path = Path(target)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path: Path | None = None

    try:
        with tempfile.NamedTemporaryFile(
            "w",
            encoding="utf-8",
            newline="\n",
            dir=target_path.parent,
            prefix=f".{target_path.name}.",
            suffix=".tmp",
            delete=False,
        ) as tmp:
            tmp_path = Path(tmp.name)
            result = convert_spreadsheet(source_path, tmp, options=options, display_name=display_name)
        os.replace(tmp_path, target_path)
        return result
    finally:
        if tmp_path and tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def convert_spreadsheet(
    source: str | Path,
    output: TextIO,
    options: SpreadsheetConversionOptions | None = None,
    display_name: str | None = None,
) -> SpreadsheetConversionResult:
    """Stream a spreadsheet to a Markdown writer and return bounded preview metadata."""
    options = options or SpreadsheetConversionOptions()
    source_path = Path(source)
    extension = source_path.suffix.lower()
    if extension not in SPREADSHEET_EXTENSIONS:
        raise SpreadsheetConversionError(f"Unsupported spreadsheet extension: {extension}")

    writer = _LimitedWriter(output, options.max_output_chars)
    preview = _PreviewBuilder(options.preview_max_chars)
    state = _ConversionState(started_at=time.monotonic())

    if extension == ".csv":
        _convert_csv(source_path, writer, preview, options, state, display_name)
    elif extension == ".xlsx":
        _convert_xlsx(source_path, writer, preview, options, state)
    else:
        _convert_xls(source_path, writer, preview, options, state)

    if state.cell_truncated:
        state.warnings.append(
            f"One or more cells exceeded {options.max_cell_chars} characters and were truncated."
        )

    return SpreadsheetConversionResult(
        markdown_length=writer.chars,
        preview=preview.text(),
        preview_truncated=preview.truncated,
        sheets=state.sheets,
        rows=state.rows,
        warnings=state.warnings,
    )


def _convert_csv(
    source: Path,
    writer: _LimitedWriter,
    preview: _PreviewBuilder,
    options: SpreadsheetConversionOptions,
    state: _ConversionState,
    display_name: str | None,
) -> None:
    encoding, dialect = _detect_csv_format(source)
    sheet_name = Path(display_name).stem if display_name else source.stem
    with source.open("r", encoding=encoding, errors="replace", newline="") as handle:
        reader = csv.reader(handle, dialect)
        _convert_sheet(sheet_name or "CSV", reader, writer, preview, options, state)


def _convert_xlsx(
    source: Path,
    writer: _LimitedWriter,
    preview: _PreviewBuilder,
    options: SpreadsheetConversionOptions,
    state: _ConversionState,
) -> None:
    _check_xlsx_archive(source, options)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SpreadsheetConversionError("XLSX conversion requires openpyxl.") from exc

    workbook = load_workbook(filename=source, read_only=True, data_only=True)
    try:
        worksheets = workbook.worksheets
        if len(worksheets) > options.max_sheets:
            raise SpreadsheetLimitError(
                f"Workbook has {len(worksheets)} sheets; maximum is {options.max_sheets}."
            )
        for worksheet in worksheets:
            _convert_sheet(
                worksheet.title,
                worksheet.iter_rows(values_only=True),
                writer,
                preview,
                options,
                state,
            )
    finally:
        workbook.close()


def _convert_xls(
    source: Path,
    writer: _LimitedWriter,
    preview: _PreviewBuilder,
    options: SpreadsheetConversionOptions,
    state: _ConversionState,
) -> None:
    try:
        import xlrd
    except ImportError as exc:
        raise SpreadsheetConversionError("XLS conversion requires xlrd.") from exc

    workbook = xlrd.open_workbook(str(source), on_demand=True)
    try:
        if workbook.nsheets > options.max_sheets:
            raise SpreadsheetLimitError(
                f"Workbook has {workbook.nsheets} sheets; maximum is {options.max_sheets}."
            )
        for index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(index)
            _convert_sheet(
                sheet.name,
                _iter_xls_rows(sheet, workbook.datemode, xlrd),
                writer,
                preview,
                options,
                state,
            )
    finally:
        workbook.release_resources()


def _convert_sheet(
    sheet_name: str,
    rows: Iterable[Sequence[object]],
    writer: _LimitedWriter,
    preview: _PreviewBuilder,
    options: SpreadsheetConversionOptions,
    state: _ConversionState,
) -> None:
    state.sheets += 1
    if state.sheets > options.max_sheets:
        raise SpreadsheetLimitError(f"Maximum sheet count exceeded: {options.max_sheets}.")

    title = sheet_name or f"Sheet {state.sheets}"
    heading = f"## {_escape_heading(title)}\n\n"
    writer.write(heading)
    preview.write(heading)

    iterator = iter(rows)
    buffered_rows: list[list[str]] = []
    for row in iterator:
        _check_timeout(options, state)
        normalized = _normalize_row(row, options, state)
        if options.skip_empty_rows and _is_empty_row(normalized):
            continue
        buffered_rows.append(normalized)
        if len(buffered_rows) >= options.header_scan_rows:
            break

    if not buffered_rows:
        empty_text = "_Empty sheet._\n\n"
        writer.write(empty_text)
        preview.write(empty_text)
        return

    width = max(len(row) for row in buffered_rows)
    _check_width(title, width, options)

    header = _prepare_header(buffered_rows[0], width)
    _write_table_header(header, writer, preview)
    state.rows += 1

    preview_data_rows = 0
    preview_row_limit_marked = False

    for row in buffered_rows[1:]:
        preview_data_rows, preview_row_limit_marked = _write_data_row(
            title,
            row,
            width,
            preview_data_rows,
            preview_row_limit_marked,
            writer,
            preview,
            options,
            state,
        )

    for row in iterator:
        _check_timeout(options, state)
        normalized = _normalize_row(row, options, state)
        if options.skip_empty_rows and _is_empty_row(normalized):
            continue
        preview_data_rows, preview_row_limit_marked = _write_data_row(
            title,
            normalized,
            width,
            preview_data_rows,
            preview_row_limit_marked,
            writer,
            preview,
            options,
            state,
        )

    writer.write("\n")
    preview.write("\n")


def _write_table_header(
    header: Sequence[str],
    writer: _LimitedWriter,
    preview: _PreviewBuilder,
) -> None:
    header_line = _markdown_row(header)
    separator = _markdown_row(["---"] * len(header))
    writer.write(header_line)
    writer.write(separator)
    preview.write(header_line)
    preview.write(separator)


def _write_data_row(
    sheet_name: str,
    row: Sequence[str],
    width: int,
    preview_data_rows: int,
    preview_row_limit_marked: bool,
    writer: _LimitedWriter,
    preview: _PreviewBuilder,
    options: SpreadsheetConversionOptions,
    state: _ConversionState,
) -> tuple[int, bool]:
    fitted_row = _fit_row(sheet_name, row, width, state)
    row_line = _markdown_row(fitted_row)
    writer.write(row_line)
    state.rows += 1

    if preview_data_rows < options.preview_rows_per_sheet:
        preview.write(row_line)
        return preview_data_rows + 1, preview_row_limit_marked

    if not preview_row_limit_marked:
        preview.add_marker(
            f"Preview truncated after {options.preview_rows_per_sheet} data rows "
            f"for sheet '{sheet_name}'. Download the full Markdown file for complete output."
        )
        return preview_data_rows, True

    return preview_data_rows, preview_row_limit_marked


def _normalize_row(
    row: Sequence[object],
    options: SpreadsheetConversionOptions,
    state: _ConversionState,
) -> list[str]:
    normalized = [_format_cell(value, options, state) for value in row]
    while normalized and normalized[-1] == "":
        normalized.pop()
    return normalized


def _format_cell(
    value: object,
    options: SpreadsheetConversionOptions,
    state: _ConversionState,
) -> str:
    if value is None:
        text = ""
    elif isinstance(value, datetime):
        text = value.isoformat(sep=" ", timespec="seconds")
    elif isinstance(value, date):
        text = value.isoformat()
    elif isinstance(value, datetime_time):
        text = value.isoformat(timespec="seconds")
    else:
        text = str(value)

    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    text = text.replace("|", r"\|").strip()

    if len(text) > options.max_cell_chars:
        state.cell_truncated = True
        text = text[: options.max_cell_chars].rstrip() + "... [truncated]"

    return text


def _prepare_header(row: Sequence[str], width: int) -> list[str]:
    header = _fit_cells(row, width)
    return [cell if cell else f"Column {index + 1}" for index, cell in enumerate(header)]


def _fit_row(
    sheet_name: str,
    row: Sequence[str],
    width: int,
    state: _ConversionState,
) -> list[str]:
    if len(row) > width and sheet_name not in state.row_width_warning_sheets:
        state.warnings.append(
            f"Sheet '{sheet_name}' had rows wider than the detected header; extra cells were omitted."
        )
        state.row_width_warning_sheets.add(sheet_name)
    return _fit_cells(row, width)


def _fit_cells(row: Sequence[str], width: int) -> list[str]:
    fitted = list(row[:width])
    if len(fitted) < width:
        fitted.extend([""] * (width - len(fitted)))
    return fitted


def _markdown_row(cells: Sequence[str]) -> str:
    return "| " + " | ".join(cells) + " |\n"


def _is_empty_row(row: Sequence[str]) -> bool:
    return all(cell == "" for cell in row)


def _check_width(sheet_name: str, width: int, options: SpreadsheetConversionOptions) -> None:
    if width < 1:
        return
    if width > options.max_columns:
        raise SpreadsheetLimitError(
            f"Sheet '{sheet_name}' has {width} columns; maximum is {options.max_columns}."
        )


def _check_timeout(options: SpreadsheetConversionOptions, state: _ConversionState) -> None:
    if time.monotonic() - state.started_at > options.timeout_seconds:
        raise SpreadsheetLimitError(
            f"Spreadsheet conversion exceeded {options.timeout_seconds} seconds."
        )


def _escape_heading(text: str) -> str:
    return text.replace("\n", " ").strip() or "Sheet"


def _detect_csv_format(source: Path) -> tuple[str, csv.Dialect]:
    with source.open("rb") as handle:
        sample = handle.read(64 * 1024)

    encoding = "utf-8-sig"
    if sample:
        detected = from_bytes(sample).best()
        if detected and detected.encoding:
            encoding = detected.encoding

    sample_text = sample.decode(encoding, errors="replace") if sample else ""
    try:
        dialect = (
            csv.Sniffer().sniff(sample_text, delimiters=",\t;|")
            if sample_text
            else csv.excel
        )
    except csv.Error:
        dialect = csv.excel
    return encoding, dialect


def _check_xlsx_archive(source: Path, options: SpreadsheetConversionOptions) -> None:
    try:
        with zipfile.ZipFile(source) as archive:
            infos = archive.infolist()
    except zipfile.BadZipFile as exc:
        raise SpreadsheetConversionError("XLSX file is not a valid ZIP archive.") from exc

    uncompressed = sum(item.file_size for item in infos)
    compressed = max(sum(item.compress_size for item in infos), 1)
    if uncompressed > options.max_xlsx_uncompressed_bytes:
        raise SpreadsheetLimitError(
            "XLSX expanded content is too large "
            f"({uncompressed} bytes; maximum is {options.max_xlsx_uncompressed_bytes})."
        )
    if uncompressed / compressed > options.max_xlsx_compression_ratio:
        raise SpreadsheetLimitError(
            "XLSX compression ratio is too high for safe processing "
            f"({uncompressed / compressed:.1f}; maximum is {options.max_xlsx_compression_ratio})."
        )


def _iter_xls_rows(sheet, datemode: int, xlrd_module) -> Iterator[list[object]]:
    for row_index in range(sheet.nrows):
        row: list[object] = []
        for col_index in range(sheet.ncols):
            ctype = sheet.cell_type(row_index, col_index)
            value = sheet.cell_value(row_index, col_index)
            if ctype == xlrd_module.XL_CELL_EMPTY:
                row.append(None)
            elif ctype == xlrd_module.XL_CELL_DATE:
                try:
                    row.append(xlrd_module.xldate_as_datetime(value, datemode))
                except Exception:
                    row.append(value)
            elif ctype == xlrd_module.XL_CELL_BOOLEAN:
                row.append(bool(value))
            elif ctype == xlrd_module.XL_CELL_ERROR:
                row.append(xlrd_module.error_text_from_code.get(value, f"#ERR{value}"))
            else:
                row.append(value)
        yield row
